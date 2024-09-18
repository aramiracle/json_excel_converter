import json
import pandas as pd
from openpyxl.styles import Alignment
from pathlib import Path

class JsonExcelConverter:
    def __init__(self, json_file=None, data_dict=None, excel_file=None, output_json_file=None):
        """
        Initialize the JsonExcelConverter with optional input and output files.

        Args:
            json_file (str, optional): Path to the input JSON file.
            data_dict (dict, optional): Dictionary to be converted to Excel.
            excel_file (str, optional): Path to the input Excel file.
            output_json_file (str, optional): Path to the output JSON file.
        """
        self.json_file = json_file
        self.data_dict = data_dict
        self.excel_file = excel_file
        self.output_json_file = output_json_file

    def json_to_excel(self):
        """
        Convert JSON data to an Excel file.

        This method requires either a JSON file or a dictionary to be provided. The output is saved
        as an Excel file. The generated Excel file will have a flattened structure, and the headers
        will reflect the hierarchical levels.

        Raises:
            ValueError: If neither json_file nor data_dict is provided, or if excel_file is not provided.
            ValueError: If JSON data is not consistent in depth.
        """
        if self.json_file:
            self.data_dict = self._load_json(self.json_file)
            self._validate_json_depth(self.data_dict)
            self.excel_file = self._generate_excel_filename(self.json_file)
        elif self.data_dict is None or self.excel_file is None:
            raise ValueError("Both json_file (or data_dict) and excel_file must be provided for JSON to Excel conversion.")

        data_frames = self._dict_to_dataframe(self.data_dict)
        self._save_to_excel(data_frames, self.excel_file)

    def excel_to_json(self):
        """
        Convert Excel data to a JSON file.

        This method requires both an Excel file and an output JSON file to be provided. The Excel file
        must have a flattened structure where all JSON data is at the same depth. The output will be
        a nested dictionary saved as a JSON file.

        Raises:
            ValueError: If either excel_file or output_json_file is not provided.
            ValueError: If Excel data is not consistent in depth.
        """
        if not self.excel_file or not self.output_json_file:
            raise ValueError("Both excel_file and output_json_file must be provided for Excel to JSON conversion.")

        nested_dict = self._excel_to_dict()
        self._validate_excel_depth(nested_dict)
        self._save_to_json(nested_dict)

    def _load_json(self, json_file):
        """
        Load JSON data from a file.

        Args:
            json_file (str): Path to the JSON file.

        Returns:
            dict: Parsed JSON data.
        """
        with open(json_file, 'r') as f:
            return json.load(f)

    def _generate_excel_filename(self, json_file):
        """
        Generate an Excel filename based on the JSON filename.

        Args:
            json_file (str): Path to the JSON file.

        Returns:
            str: Generated Excel filename.
        """
        base = Path(json_file).stem
        return f'{base}.xlsx'

    def _dict_to_dataframe(self, data_dict):
        """
        Convert a dictionary to a DataFrame with a flattened structure.

        Args:
            data_dict (dict): Dictionary to be converted.

        Returns:
            pd.DataFrame: Flattened DataFrame.
        """
        rows = self._flatten_dict(data_dict)
        max_depth = max(len(row) for row in rows)
        return pd.DataFrame(rows, columns=[f'Level_{i+1}' for i in range(max_depth)])

    def _flatten_dict(self, d, parent_key=()):
        """
        Flatten a nested dictionary into a list of tuples.

        Args:
            d (dict): Dictionary to be flattened.
            parent_key (tuple, optional): Key tuple for nesting.

        Returns:
            list: List of tuples representing flattened dictionary.
        """
        rows = []
        for k, v in d.items():
            if isinstance(v, dict):
                rows.extend(self._flatten_dict(v, parent_key + (k,)))
            elif isinstance(v, list) and v:  # Only process non-empty lists
                for item in v:
                    rows.append(parent_key + (k, item))
            else:
                rows.append(parent_key + (k, v))
        return rows

    def _save_to_excel(self, df, excel_file):
        """
        Save the DataFrame to an Excel file with formatting.

        Args:
            df (pd.DataFrame): DataFrame to be saved.
            excel_file (str): Path to the output Excel file.
        """
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, startrow=1)
            worksheet = writer.sheets['Sheet1']

            self._set_headers(worksheet, df)
            self._merge_cells(worksheet)

    def _set_headers(self, ws, df):
        """
        Set headers and format cells in the Excel sheet.

        Args:
            ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet object.
            df (pd.DataFrame): DataFrame used for header setting.
        """
        headers = [f'Level {i+1}' for i in range(df.shape[1])]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _merge_cells(self, ws):
        """
        Merge cells with the same value in each column.

        Args:
            ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet object.
        """
        for col in range(ws.max_column):
            col_letter = ws.cell(row=2, column=col+1).column_letter
            start_row = 2
            prev_value = None
            for row in range(2, ws.max_row+1):
                cell_value = ws.cell(row=row, column=col+1).value
                if cell_value != prev_value:
                    if start_row != row - 1 and prev_value is not None:
                        ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{row-1}')
                        ws.cell(row=start_row, column=col+1).alignment = Alignment(vertical='center')
                    prev_value = cell_value
                    start_row = row
            if start_row != ws.max_row and prev_value is not None:
                ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{ws.max_row}')
                ws.cell(row=start_row, column=col+1).alignment = Alignment(vertical='center')

    def _excel_to_dict(self):
        """
        Convert Excel data to a nested dictionary.

        Returns:
            dict: Nested dictionary representation of the Excel data.
        """
        df = pd.read_excel(self.excel_file, header=None).iloc[1:].reset_index(drop=True)
        df.ffill(axis=0, inplace=True)
        return self._build_dict(df)

    def _build_dict(self, df):
        """
        Build a nested dictionary from the DataFrame.

        Args:
            df (pd.DataFrame): DataFrame to be converted.

        Returns:
            dict: Nested dictionary built from the DataFrame.
        """
        result = {}
        for _, row in df.iterrows():
            row = row.tolist()
            current_level = result
            for i in range(len(row) - 2):
                value = row[i]
                if pd.notna(value):
                    if value not in current_level:
                        current_level[value] = {}
                    current_level = current_level[value]
            last_key, last_value = row[-2], row[-1]
            if pd.notna(last_key):
                if last_key not in current_level:
                    current_level[last_key] = []
                if pd.notna(last_value) and last_value not in current_level[last_key]:
                    current_level[last_key].append(last_value)
        return result

    def _save_to_json(self, nested_dict):
        """
        Save the nested dictionary to a JSON file.

        Args:
            nested_dict (dict): Nested dictionary to be saved.
        """
        with open(self.output_json_file, 'w') as f:
            json.dump(nested_dict, f, indent=4)

    def _validate_json_depth(self, data):
        """
        Validate that all JSON data is at the same depth.

        Args:
            data (dict): JSON data to be validated.

        Raises:
            ValueError: If JSON data is not consistent in depth.
        """
        def get_depth(d, current_depth=0):
            if isinstance(d, dict):
                if not d:
                    return current_depth
                return max(get_depth(v, current_depth + 1) for v in d.values())
            elif isinstance(d, list):
                if not d:
                    return current_depth
                return max(get_depth(item, current_depth) for item in d)
            else:
                return current_depth

        max_depth = get_depth(data)
        
        def check_depth(d, depth=0):
            if isinstance(d, dict):
                if any(get_depth(v, depth + 1) != max_depth for v in d.values()):
                    return False
                return all(check_depth(v, depth + 1) for v in d.values())
            elif isinstance(d, list):
                if any(get_depth(item, depth) != max_depth for item in d):
                    return False
                return all(check_depth(item, depth) for item in d)
            return True

        if not check_depth(data):
            raise ValueError("All JSON data must be at the same depth.")
    
    def _validate_excel_depth(self, data):
        """
        Validate that all rows in the Excel data have the same depth.

        Args:
            data (dict): Nested dictionary to be validated.

        Raises:
            ValueError: If Excel data is not consistent in depth.
        """
        def get_row_depth(row):
            return len([item for item in row if pd.notna(item)])

        def check_depth(df):
            row_depths = df.apply(lambda row: get_row_depth(row.tolist()), axis=1)
            min_depth = row_depths.min()
            if row_depths.max() != min_depth:
                return False
            return True

        df = pd.DataFrame.from_dict(data, orient='index')
        if not check_depth(df):
            raise ValueError("All Excel rows must have the same depth.")
